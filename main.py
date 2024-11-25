import glob
import re
from pathlib import Path

import openpyxl
from pony.orm import db_session

from boomer_utils import parse_yes_or_no, re_contains_words, throw_error
from generate_reports import generate_master_report, generate_participants_sheet, generate_judge_report
from models import School, Participant, Event, Registration


@db_session
def main():
    registration_files = [f for f in glob.glob("*.xlsx") if not ("template" in f.lower())]
    
    if not registration_files:
        print("NO VALID REGISTRATION FILES FOUND")
        return

    enable_participants_sheets = input("GENERATE LISTS OF UNIQUE STUDENT NAMES PER SCHOOL? (YES/NO) ")
    print(f"FOUND {len(registration_files)} REGISTRATION FILE(S)")
    event_count = import_events(registration_files[0])  # Import event listings from the first workbook
    print(f"IMPORTED {event_count} EVENTS")

    schools = []
    for workbook_file in registration_files:
        
        
        workbook = openpyxl.load_workbook(workbook_file)
        worksheet = workbook['Original']
        school_name = worksheet.cell(4, 2).value
        print("Processing", school_name, "registration")

        row_indeces = get_row_indeces(worksheet) # Get row indeces for School model
        
        school = School(
            name=school_name,
            regular_registrations=int(worksheet.cell(row_indeces["regular_registrations"], 2).value or 0),  # Default to zero if not specified
            late_registrations=int(worksheet.cell(row_indeces["late_registrations"], 2).value or 0),
            total_enrolled=int(worksheet.cell(row_indeces["total_enrolled"], 2).value or throw_error(f"{school_name} Registration sheet does not contain a valid total number of enrolled students")),
            rookie_teacher=parse_yes_or_no(worksheet.cell(row_indeces["rookie_teacher"], 2).value or 'no'),  # Default to no if not specified
            rookie_school=parse_yes_or_no(worksheet.cell(row_indeces["rookie_school"], 2).value or 'no'),
            attending_state=parse_yes_or_no(worksheet.cell(row_indeces["attending_state"], 2).value or 'no')
        )
        schools.append(school)

        event_row = row_indeces["EVENT_START_ROW"]
        for event in Event.select():
            for _ in range(max(event.max_groups, 1)):
                participant_row = 0
                participants = []
                while participant_row < event.max_participants:
                    participant_name = worksheet.cell(event_row + participant_row, 2).value
                    if participant_name is not None and participant_name.strip():  # If a non-whitespace entry exists
                        participant = find_or_create_participant(participant_name, school)
                        participants.append(participant)
                    participant_row += 1
                if len(participants) > 0:  # Only create a registration if there are more than zero participants
                    if event.max_groups == 0:  # Individual events have a max_groups of 0
                        for participant in participants:  # Register each participant separately for individual events
                            Registration(event=event, school=school, participants=participant)
                    else:  # Register participants together for group events
                        Registration(event=event, school=school, participants=participants)
                event_row += participant_row

    Path('output').mkdir(exist_ok=True)
    events = Event.select().order_by(Event.name)
    schools = School.select().order_by(School.name)
    generate_master_report(events, schools)
    generate_judge_report(events)
    
    if parse_yes_or_no(enable_participants_sheets):
        Path('output/student names by school').mkdir(exist_ok=True)
        for school in schools:
            generate_participants_sheet(school)

    print()
    print("SCRIPT WRITTEN BY DAMIAN LALL, CHS '21")
    print("REVISED AND UPDATED BY TIMOTHY LIU, CHS '25;")
    print("                       ANSHUL MAGO, CHS '25;")
    print("                       RAGHAV KENCHANNAVAR, CHS '25")
    print()
    print("TASK COMPLETE, PRESS ENTER TO EXIT", end='')
    input()


def import_events(workbook_file) -> int:
    workbook = openpyxl.load_workbook(workbook_file)
    worksheet = workbook['Original']
    participant_count = 0
    previous_event = None
    
    # Find start index
    for row in worksheet.iter_rows():
        row_label = row[0].value
        row_index = row[0].row
        if re_contains_words(["Event",], row_label):
            event_start_row = row_index + 1
    
    for row in worksheet.iter_rows(event_start_row):
        row_event = row[0].value
        if previous_event is not None and row_event != previous_event:
            create_event(previous_event, participant_count)
            participant_count = 0
        participant_count += 1
        previous_event = row_event
    create_event(previous_event, participant_count)

    return Event.select().count()


def get_row_indeces(worksheet):
    row_indeces = {
        "regular_registrations": 16,
        "late_registrations": 17,
        "total_enrolled": 19,
        "rookie_teacher": 20,
        "rookie_school": 21,
        "attending_state": 22,
        "EVENT_START_ROW": 25,
    }
    for row in worksheet.iter_rows():
        row_label = row[0].value
        row_index = row[0].row
        
        if re_contains_words(['Number', 'regular'], row_label):
            row_indeces["regular_registrations"] = row_index
        elif re_contains_words(['Number', 'late'], row_label):
            row_indeces["late_registrations"] = row_index
        elif re_contains_words(['Total', 'enroll'], row_label):
            row_indeces["total_enrolled"] = row_index
        elif re_contains_words(['rookie', 'teacher', 'entered'], row_label):
            row_indeces["rookie_teacher"] = row_index
        elif re_contains_words(['rookie', 'school', 'entered'], row_label):
            row_indeces["rookie_school"] = row_index
        elif re_contains_words(['attend', 'State'], row_label):
            row_indeces["attending_state"] = row_index
        elif re_contains_words(["Event",], row_label):
            row_indeces["EVENT_START_ROW"] = row_index+1
            
    return row_indeces


def create_event(event_name, participant_count):
    is_group = "Group" in event_name
    event_name = re.sub(R"[(\[].*?[)\]]", "", event_name).strip()  # Remove anything between parenthesis
    event = Event.get(name=event_name)
    if event is not None:
        event.max_groups += 1
    else:
        Event(name=event_name, max_participants=participant_count, max_groups=is_group)


def find_or_create_participant(name, school) -> Participant:
    participant = Participant.get(name=name, school=school)
    if participant is None:
        return Participant(name=name, school=school)
    return participant


if __name__ == '__main__':
    main()
