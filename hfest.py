import glob
import re

import openpyxl
from pony.orm import db_session

from boomer_utils import parse_yes_or_no
from generate_reports import generate_judge_sheet, generate_master_report
from models import School, Participant, Event, Registration


@db_session
def main():
    registration_files = glob.glob("Reg.*.xlsx")
    print("Found files", registration_files)
    import_events(registration_files[0])
    print("Imported events")
    schools = list()
    for workbook_file in registration_files:
        workbook = openpyxl.load_workbook(workbook_file)
        worksheet = workbook['Original']
        school_name = worksheet.cell(4, 2).value
        print("Processing", school_name)

        school = School(
            name=school_name,
            regular_registrations=worksheet.cell(16, 2).value,
            late_registrations=worksheet.cell(17, 2).value or 0,
            total_enrolled=worksheet.cell(19, 2).value,
            rookie_teacher=parse_yes_or_no(worksheet.cell(20, 2).value),
            rookie_school=parse_yes_or_no(worksheet.cell(21, 2).value),
            attending_state=parse_yes_or_no(worksheet.cell(22, 2).value)
        )
        schools.append(school)

        current_row = 37

        for event in Event.select():
            for group in range(max(event.max_groups, 1)):
                current_participant = 0
                participants = list()
                while current_participant < event.max_participants:
                    participant_name = worksheet.cell(current_row + current_participant, 2).value
                    if participant_name is not None:
                        participant = find_or_create_participant(participant_name, school)
                        participants.append(participant)
                    current_participant += 1
                if len(participants) > 0:
                    Registration(event=event, participants=participants)
                current_row += current_participant

    generate_master_report()
    generate_judge_sheet()


def import_events(workbook_file):
    workbook = openpyxl.load_workbook(workbook_file)
    worksheet = workbook['Original']
    participant_count = 0
    previous_event = None
    for row in worksheet.iter_rows(37):
        row_event = row[0].value
        if previous_event is not None and row_event != previous_event:
            create_event(previous_event, participant_count)
            participant_count = 0
        participant_count += 1
        previous_event = row_event
    create_event(previous_event, participant_count)


def create_event(event_name, participant_count):
    is_group = "Group" in event_name
    event_name = re.sub(R"[(\[].*?[)\]]", "", event_name).strip()
    if Event.get(name=event_name) is not None:
        event = Event.get(name=event_name)  # TODO: Use assignment operator when upgraded to 3.8
        event.max_groups += 1
    else:
        Event(name=event_name, max_participants=participant_count, max_groups=is_group)


def find_or_create_participant(name, school) -> Participant:
    participant = Participant.get(name=name, school=school)
    if participant is None:
        return Participant(name=name, school=school)
    else:
        return participant


if __name__ == '__main__':
    main()
